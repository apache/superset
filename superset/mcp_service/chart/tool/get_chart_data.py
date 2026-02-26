# Licensed to the Apache Software Foundation (ASF) under one
# or more contributor license agreements.  See the NOTICE file
# distributed with this work for additional information
# regarding copyright ownership.  The ASF licenses this file
# to you under the Apache License, Version 2.0 (the
# "License"); you may not use this file except in compliance
# with the License.  You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
# KIND, either express or implied.  See the License for the
# specific language governing permissions and limitations
# under the License.

"""
MCP tool: get_chart_data
"""

import logging
import time
from typing import Any, Dict, List, TYPE_CHECKING

from fastmcp import Context
from flask import current_app
from superset_core.mcp import tool

if TYPE_CHECKING:
    from superset.models.slice import Slice

from superset.commands.exceptions import CommandException
from superset.commands.explore.form_data.parameters import CommandParameters
from superset.exceptions import SupersetException
from superset.extensions import event_logger
from superset.mcp_service.chart.chart_utils import validate_chart_dataset
from superset.mcp_service.chart.schemas import (
    ChartData,
    ChartError,
    DataColumn,
    GetChartDataRequest,
    PerformanceMetadata,
)
from superset.mcp_service.utils.cache_utils import get_cache_status_from_result
from superset.mcp_service.utils.schema_utils import parse_request

logger = logging.getLogger(__name__)


def _get_cached_form_data(form_data_key: str) -> str | None:
    """Retrieve form_data from cache using form_data_key.

    Returns the JSON string of form_data if found, None otherwise.
    """
    from superset.commands.explore.form_data.get import GetFormDataCommand

    try:
        cmd_params = CommandParameters(key=form_data_key)
        return GetFormDataCommand(cmd_params).run()
    except (KeyError, ValueError, CommandException) as e:
        logger.warning("Failed to retrieve form_data from cache: %s", e)
        return None


@tool(tags=["data"])
@parse_request(GetChartDataRequest)
async def get_chart_data(  # noqa: C901
    request: GetChartDataRequest, ctx: Context
) -> ChartData | ChartError:
    """Get chart data by ID or UUID.

    Returns the actual data behind a chart for LLM analysis without image rendering.

    Supports:
    - Numeric ID or UUID lookup
    - Multiple formats: json, csv, excel
    - Cache control: use_cache, force_refresh, cache_timeout
    - Optional row limit override (respects chart's configured limits)
    - form_data_key: retrieves data using unsaved chart configuration from Explore

    When form_data_key is provided, the tool uses the cached (unsaved) chart
    configuration to query data, allowing you to get data for what the user
    actually sees in the Explore view (not the saved version).

    Returns underlying data in requested format with cache status.
    """
    await ctx.info(
        "Starting chart data retrieval: identifier=%s, format=%s, limit=%s, "
        "form_data_key=%s"
        % (
            request.identifier,
            request.format,
            request.limit,
            request.form_data_key,
        )
    )
    await ctx.debug(
        "Cache settings: use_cache=%s, force_refresh=%s, cache_timeout=%s"
        % (
            request.use_cache,
            request.force_refresh,
            request.cache_timeout,
        )
    )

    try:
        await ctx.report_progress(1, 4, "Looking up chart")
        from superset.daos.chart import ChartDAO
        from superset.utils import json as utils_json

        # Find the chart
        with event_logger.log_context(action="mcp.get_chart_data.chart_lookup"):
            chart = None
            if isinstance(request.identifier, int) or (
                isinstance(request.identifier, str) and request.identifier.isdigit()
            ):
                chart_id = (
                    int(request.identifier)
                    if isinstance(request.identifier, str)
                    else request.identifier
                )
                await ctx.debug(
                    "Performing ID-based chart lookup: chart_id=%s" % (chart_id,)
                )
                chart = ChartDAO.find_by_id(chart_id)
            else:
                await ctx.debug(
                    "Performing UUID-based chart lookup: uuid=%s"
                    % (request.identifier,)
                )
                # Try UUID lookup using DAO flexible method
                chart = ChartDAO.find_by_id(request.identifier, id_column="uuid")

        if not chart:
            await ctx.error("Chart not found: identifier=%s" % (request.identifier,))
            return ChartError(
                error=f"No chart found with identifier: {request.identifier}",
                error_type="NotFound",
            )

        await ctx.info(
            "Chart found successfully: chart_id=%s, chart_name=%s, viz_type=%s"
            % (
                chart.id,
                chart.slice_name,
                chart.viz_type,
            )
        )
        logger.info("Getting data for chart %s: %s", chart.id, chart.slice_name)

        # Validate the chart's dataset is accessible before retrieving data
        validation_result = validate_chart_dataset(chart, check_access=True)
        if not validation_result.is_valid:
            await ctx.warning(
                "Chart found but dataset is not accessible: %s"
                % (validation_result.error,)
            )
            return ChartError(
                error=validation_result.error
                or "Chart's dataset is not accessible. Dataset may have been deleted.",
                error_type="DatasetNotAccessible",
            )
        # Log any warnings (e.g., virtual dataset warnings)
        for warning in validation_result.warnings:
            await ctx.warning("Dataset warning: %s" % (warning,))

        start_time = time.time()

        # Track whether we're using unsaved state
        using_unsaved_state = False
        cached_form_data_dict = None

        try:
            await ctx.report_progress(2, 4, "Preparing data query")
            from superset.charts.schemas import ChartDataQueryContextSchema
            from superset.commands.chart.data.get_data_command import ChartDataCommand

            # Check if form_data_key is provided - use cached form_data instead
            if request.form_data_key:
                await ctx.info(
                    "Retrieving unsaved chart state from cache: form_data_key=%s"
                    % (request.form_data_key,)
                )
                if cached_form_data := _get_cached_form_data(request.form_data_key):
                    try:
                        parsed_form_data = utils_json.loads(cached_form_data)
                        # Only use if it's actually a dict (not null, list, etc.)
                        if isinstance(parsed_form_data, dict):
                            cached_form_data_dict = parsed_form_data
                            using_unsaved_state = True
                            await ctx.info(
                                "Using cached form_data from form_data_key "
                                "for data query"
                            )
                        else:
                            await ctx.warning(
                                "Cached form_data is not a JSON object. "
                                "Falling back to saved chart configuration."
                            )
                    except (TypeError, ValueError) as e:
                        await ctx.warning(
                            "Failed to parse cached form_data: %s. "
                            "Falling back to saved chart configuration." % str(e)
                        )
                else:
                    await ctx.warning(
                        "form_data_key provided but no cached data found. "
                        "The cache may have expired. Using saved chart configuration."
                    )

            # Use the chart's saved query_context - this is the key!
            # The query_context contains all the information needed to reproduce
            # the chart's data exactly as shown in the visualization
            query_context_json = None

            # If using cached form_data, we need to build query_context from it
            if using_unsaved_state and cached_form_data_dict is not None:
                # Build query context from cached form_data (unsaved state)
                from superset.common.query_context_factory import QueryContextFactory

                factory = QueryContextFactory()
                row_limit = (
                    request.limit
                    or cached_form_data_dict.get("row_limit")
                    or current_app.config["ROW_LIMIT"]
                )

                # Get datasource info from cached form_data or fall back to chart
                datasource_id = cached_form_data_dict.get(
                    "datasource_id", chart.datasource_id
                )
                datasource_type = cached_form_data_dict.get(
                    "datasource_type", chart.datasource_type
                )

                # Handle different chart types that have different form_data
                # structures. Some charts use "metric" (singular), not "metrics"
                # (plural): big_number, big_number_total, pop_kpi.
                # These charts also don't have groupby columns.
                cached_viz_type = cached_form_data_dict.get(
                    "viz_type", chart.viz_type or ""
                )
                if cached_viz_type in ("big_number", "big_number_total", "pop_kpi"):
                    metric = cached_form_data_dict.get("metric")
                    cached_metrics = [metric] if metric else []
                    cached_groupby: list[str] = []
                else:
                    cached_metrics = cached_form_data_dict.get("metrics", [])
                    cached_groupby = cached_form_data_dict.get("groupby", [])

                query_context = factory.create(
                    datasource={
                        "id": datasource_id,
                        "type": datasource_type,
                    },
                    queries=[
                        {
                            "filters": cached_form_data_dict.get("filters", []),
                            "columns": cached_groupby,
                            "metrics": cached_metrics,
                            "row_limit": row_limit,
                            "order_desc": cached_form_data_dict.get("order_desc", True),
                        }
                    ],
                    form_data=cached_form_data_dict,
                    force=request.force_refresh,
                )
                await ctx.debug(
                    "Built query_context from cached form_data (unsaved state)"
                )
            elif chart.query_context:
                try:
                    query_context_json = utils_json.loads(chart.query_context)
                    await ctx.debug(
                        "Using chart's saved query_context for data retrieval"
                    )
                except (TypeError, ValueError) as e:
                    await ctx.warning(
                        "Failed to parse chart query_context: %s" % str(e)
                    )

            if query_context_json is None and not using_unsaved_state:
                # Fallback: Chart has no saved query_context
                # This can happen with older charts that haven't been re-saved
                await ctx.warning(
                    "Chart has no saved query_context. "
                    "Data may not match the chart visualization exactly. "
                    "Consider re-saving the chart to enable full data retrieval."
                )
                # Try to construct from form_data as a fallback
                form_data = utils_json.loads(chart.params) if chart.params else {}
                from superset.common.query_context_factory import QueryContextFactory

                factory = QueryContextFactory()
                row_limit = (
                    request.limit
                    or form_data.get("row_limit")
                    or current_app.config["ROW_LIMIT"]
                )

                # Handle different chart types that have different form_data
                # structures.  Chart types that exclusively use "metric"
                # (singular) with no groupby:
                #   big_number, big_number_total, pop_kpi
                # Chart types that use "metric" (singular) but may have
                # groupby-like fields (entity, series, columns):
                #   world_map, treemap_v2, sunburst_v2, gauge_chart
                # Bubble charts use x/y/size as separate metric fields.
                viz_type = chart.viz_type or ""

                singular_metric_no_groupby = (
                    "big_number",
                    "big_number_total",
                    "pop_kpi",
                )
                singular_metric_types = (
                    *singular_metric_no_groupby,
                    "world_map",
                    "treemap_v2",
                    "sunburst_v2",
                    "gauge_chart",
                )

                if viz_type == "bubble":
                    # Bubble charts store metrics in x, y, size fields
                    bubble_metrics = []
                    for field in ("x", "y", "size"):
                        m = form_data.get(field)
                        if m:
                            bubble_metrics.append(m)
                    metrics = bubble_metrics
                    groupby_columns: list[str] = list(
                        form_data.get("entity", None) and [form_data["entity"]] or []
                    )
                    series_field = form_data.get("series")
                    if series_field and series_field not in groupby_columns:
                        groupby_columns.append(series_field)
                elif viz_type in singular_metric_types:
                    # These chart types use "metric" (singular)
                    metric = form_data.get("metric")
                    metrics = [metric] if metric else []
                    if viz_type in singular_metric_no_groupby:
                        groupby_columns = []
                    else:
                        # Some singular-metric charts use groupby, entity,
                        # series, or columns for dimensional breakdown
                        groupby_columns = list(form_data.get("groupby") or [])
                        entity = form_data.get("entity")
                        if entity and entity not in groupby_columns:
                            groupby_columns.append(entity)
                        series = form_data.get("series")
                        if series and series not in groupby_columns:
                            groupby_columns.append(series)
                        form_columns = form_data.get("columns")
                        if form_columns and isinstance(form_columns, list):
                            for col in form_columns:
                                if isinstance(col, str) and col not in groupby_columns:
                                    groupby_columns.append(col)
                else:
                    # Standard charts use "metrics" (plural) and "groupby"
                    metrics = form_data.get("metrics", [])
                    groupby_columns = list(form_data.get("groupby") or [])
                    # Some chart types use "columns" instead of "groupby"
                    if not groupby_columns:
                        form_columns = form_data.get("columns")
                        if form_columns and isinstance(form_columns, list):
                            for col in form_columns:
                                if isinstance(col, str):
                                    groupby_columns.append(col)

                # Fallback: if metrics is still empty, try singular "metric"
                if not metrics:
                    fallback_metric = form_data.get("metric")
                    if fallback_metric:
                        metrics = [fallback_metric]

                # Fallback: try entity/series if groupby is still empty
                if not groupby_columns:
                    entity = form_data.get("entity")
                    if entity:
                        groupby_columns.append(entity)
                    series = form_data.get("series")
                    if series and series not in groupby_columns:
                        groupby_columns.append(series)

                # Build query columns list: include both x_axis and groupby
                x_axis_config = form_data.get("x_axis")
                query_columns = groupby_columns.copy()
                if x_axis_config and isinstance(x_axis_config, str):
                    if x_axis_config not in query_columns:
                        query_columns.insert(0, x_axis_config)
                elif x_axis_config and isinstance(x_axis_config, dict):
                    col_name = x_axis_config.get("column_name")
                    if col_name and col_name not in query_columns:
                        query_columns.insert(0, col_name)

                # Safety net: if we could not extract any metrics or
                # columns, return a clear error instead of the cryptic
                # "Empty query?" that comes from deeper in the stack.
                if not metrics and not query_columns:
                    await ctx.error(
                        "Cannot construct fallback query for chart %s "
                        "(viz_type=%s): no metrics, columns, or groupby "
                        "could be extracted from form_data. "
                        "Re-save the chart to populate query_context."
                        % (chart.id, viz_type)
                    )
                    return ChartError(
                        error=(
                            f"Chart {chart.id} (type: {viz_type}) has no "
                            f"saved query_context and its form_data does "
                            f"not contain recognizable metrics or columns. "
                            f"Please open this chart in Superset and "
                            f"re-save it to generate a query_context."
                        ),
                        error_type="MissingQueryContext",
                    )

                query_context = factory.create(
                    datasource={
                        "id": chart.datasource_id,
                        "type": chart.datasource_type,
                    },
                    queries=[
                        {
                            "filters": form_data.get("filters", []),
                            "columns": query_columns,
                            "metrics": metrics,
                            "row_limit": row_limit,
                            "order_desc": True,
                        }
                    ],
                    form_data=form_data,
                    force=request.force_refresh,
                )
            elif query_context_json is not None:
                # Apply request overrides to the saved query_context
                query_context_json["force"] = request.force_refresh

                # Apply row limit if specified (respects chart's configured limits)
                if request.limit:
                    for query in query_context_json.get("queries", []):
                        query["row_limit"] = request.limit

                # Create QueryContext from the saved context using the schema
                # This is exactly how the API does it
                query_context = ChartDataQueryContextSchema().load(query_context_json)

            await ctx.report_progress(3, 4, "Executing data query")
            await ctx.debug(
                "Query execution parameters: datasource_id=%s, datasource_type=%s, "
                "row_limit=%s, force_refresh=%s"
                % (
                    chart.datasource_id,
                    chart.datasource_type,
                    request.limit or 100,
                    request.force_refresh,
                )
            )

            # Execute the query
            with event_logger.log_context(action="mcp.get_chart_data.query_execution"):
                command = ChartDataCommand(query_context)
                result = command.run()

            # Handle empty query results for certain chart types
            if not result or ("queries" not in result) or len(result["queries"]) == 0:
                await ctx.warning(
                    "Empty query results: chart_id=%s, chart_type=%s"
                    % (chart.id, chart.viz_type)
                )
                return ChartError(
                    error=f"No query results returned for chart {chart.id}. "
                    f"This may occur with chart types like big_number.",
                    error_type="EmptyQuery",
                )

            # Extract data from result (we've already validated it exists above)
            query_result = result["queries"][0]
            data = query_result.get("data", [])
            raw_columns = query_result.get("colnames", [])

            await ctx.debug(
                "Query results received: row_count=%s, column_count=%s, "
                "has_cache_key=%s"
                % (
                    len(data),
                    len(raw_columns),
                    bool(query_result.get("cache_key")),
                )
            )

            # Check if we have data to work with
            if not data:
                await ctx.warning("No data in query results: chart_id=%s" % (chart.id,))
                return ChartError(
                    error=f"No data available for chart {chart.id}", error_type="NoData"
                )

            # Create rich column metadata
            columns = []
            for col_name in raw_columns:
                # Sample some values for metadata
                sample_values = [
                    row.get(col_name)
                    for row in data[:3]
                    if row.get(col_name) is not None
                ]

                # Infer data type
                data_type = "string"
                if sample_values:
                    if all(isinstance(v, (int, float)) for v in sample_values):
                        data_type = "numeric"
                    elif all(isinstance(v, bool) for v in sample_values):
                        data_type = "boolean"

                columns.append(
                    DataColumn(
                        name=col_name,
                        display_name=col_name.replace("_", " ").title(),
                        data_type=data_type,
                        sample_values=sample_values[:3],
                        null_count=sum(1 for row in data if row.get(col_name) is None),
                        unique_count=len({str(row.get(col_name)) for row in data}),
                    )
                )

            # Cache status information using utility function
            cache_status = get_cache_status_from_result(
                query_result, force_refresh=request.force_refresh
            )

            # Generate insights and recommendations
            insights = []
            if len(data) > 100:
                insights.append(
                    "Large dataset - consider filtering for better performance"
                )
            if len(raw_columns) > 10:
                insights.append("Many columns available - focus on key metrics")

            # Add cache-specific insights
            if cache_status.cache_hit:
                if (
                    cache_status.cache_age_seconds
                    and cache_status.cache_age_seconds > 3600
                ):
                    hours_old = cache_status.cache_age_seconds // 3600
                    insights.append(
                        f"Data is from cache ({hours_old}h old) - "
                        "consider refreshing for latest data"
                    )
                else:
                    insights.append("Data served from cache for fast response")
            else:
                insights.append("Fresh data retrieved from database")

            recommended_visualizations = []
            if any(
                "time" in col.lower() or "date" in col.lower() for col in raw_columns
            ):
                recommended_visualizations.extend(["line chart", "time series"])
            if len(raw_columns) <= 3:
                recommended_visualizations.extend(["bar chart", "scatter plot"])

            # Performance metadata with cache awareness
            execution_time = int((time.time() - start_time) * 1000)
            performance_status = (
                "cache_hit" if cache_status.cache_hit else "fresh_query"
            )
            optimization_suggestions = []

            if not cache_status.cache_hit and execution_time > 5000:
                optimization_suggestions.append(
                    "Consider using cache for this slow query"
                )
            elif (
                cache_status.cache_hit
                and cache_status.cache_age_seconds
                and cache_status.cache_age_seconds > 86400
            ):
                optimization_suggestions.append("Cache is old - consider refreshing")

            performance = PerformanceMetadata(
                query_duration_ms=execution_time,
                cache_status=performance_status,
                optimization_suggestions=optimization_suggestions,
            )

            # Generate comprehensive summary with cache info
            cache_info = ""
            if cache_status.cache_hit:
                age_info = (
                    f" (cached {cache_status.cache_age_seconds // 60}m ago)"
                    if cache_status.cache_age_seconds
                    else " (cached)"
                )
                cache_info = age_info

            summary_parts = [
                f"Chart '{chart.slice_name}' ({chart.viz_type})",
                f"Contains {len(data)} rows across {len(raw_columns)} columns"
                f"{cache_info}",
            ]

            if data and len(data) > 0:
                summary_parts.append(
                    f"Sample data includes: {', '.join(raw_columns[:3])}"
                )

            summary = ". ".join(summary_parts)

            # Handle different export formats
            if request.format == "csv":
                with event_logger.log_context(
                    action="mcp.get_chart_data.format_conversion"
                ):
                    return _export_data_as_csv(
                        chart,
                        data[: request.limit] if request.limit else data,
                        raw_columns,
                        cache_status,
                        performance,
                    )
            elif request.format == "excel":
                with event_logger.log_context(
                    action="mcp.get_chart_data.format_conversion"
                ):
                    return _export_data_as_excel(
                        chart,
                        data[: request.limit] if request.limit else data,
                        raw_columns,
                        cache_status,
                        performance,
                    )

            await ctx.report_progress(4, 4, "Building response")

            # Calculate data quality metrics
            data_completeness = 1.0 - (
                sum(col.null_count for col in columns)
                / max(len(data) * len(columns), 1)
            )

            await ctx.info(
                "Chart data retrieval completed successfully: chart_id=%s, "
                "rows_returned=%s, columns_returned=%s, execution_time_ms=%s, "
                "cache_hit=%s, data_completeness=%s"
                % (
                    chart.id,
                    len(data),
                    len(raw_columns),
                    execution_time,
                    cache_status.cache_hit,
                    round(data_completeness, 3),
                )
            )

            # Default JSON format
            return ChartData(
                chart_id=chart.id,
                chart_name=chart.slice_name or f"Chart {chart.id}",
                chart_type=chart.viz_type or "unknown",
                columns=columns,
                data=data[: request.limit] if request.limit else data,
                row_count=len(data),
                total_rows=query_result.get("rowcount"),
                summary=summary,
                insights=insights,
                data_quality={"completeness": data_completeness},
                recommended_visualizations=recommended_visualizations,
                data_freshness=None,  # Add missing field
                performance=performance,
                cache_status=cache_status,
            )

        except (CommandException, SupersetException, ValueError) as data_error:
            await ctx.error(
                "Data retrieval failed: chart_id=%s, error=%s, error_type=%s"
                % (
                    chart.id,
                    str(data_error),
                    type(data_error).__name__,
                )
            )
            logger.error("Data retrieval error for chart %s: %s", chart.id, data_error)
            return ChartError(
                error=f"Error retrieving chart data: {str(data_error)}",
                error_type="DataError",
            )

    except Exception as e:
        await ctx.error(
            "Chart data retrieval failed: identifier=%s, error=%s, error_type=%s"
            % (
                request.identifier,
                str(e),
                type(e).__name__,
            )
        )
        logger.error("Error in get_chart_data: %s", e)
        return ChartError(
            error=f"Failed to get chart data: {str(e)}", error_type="InternalError"
        )


def _export_data_as_csv(
    chart: "Slice",
    data: List[Dict[str, Any]],
    columns: List[str],
    cache_status: Any,
    performance: Any,
) -> "ChartData":
    """Export chart data as CSV format."""
    import csv
    import io

    # Create CSV content
    output = io.StringIO()

    if data and columns:
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()

        # Write data rows
        for row in data:
            # Ensure all values are properly formatted for CSV
            csv_row = {}
            for col in columns:
                value = row.get(col, "")
                # Handle None values and convert to string
                if value is None:
                    csv_row[col] = ""
                elif isinstance(value, (list, dict)):
                    csv_row[col] = str(value)
                else:
                    csv_row[col] = value
            writer.writerow(csv_row)

    csv_content = output.getvalue()

    # Return as ChartData with CSV content in a special field
    from superset.mcp_service.chart.schemas import ChartData

    return ChartData(
        chart_id=chart.id,
        chart_name=chart.slice_name or f"Chart {chart.id}",
        chart_type=chart.viz_type or "unknown",
        columns=[],  # Not needed for CSV export
        data=[],  # CSV content is in csv_data field
        row_count=len(data),
        total_rows=len(data),
        summary=f"CSV export of chart '{chart.slice_name}' with {len(data)} rows",
        insights=[f"Data exported as CSV format ({len(csv_content)} characters)"],
        data_quality={},
        recommended_visualizations=[],
        data_freshness=None,
        performance=performance,
        cache_status=cache_status,
        # Store CSV content in data field as string for the response
        csv_data=csv_content,
        format="csv",
    )


def _export_data_as_excel(
    chart: "Slice",
    data: List[Dict[str, Any]],
    columns: List[str],
    cache_status: Any,
    performance: Any,
) -> "ChartData | ChartError":
    """Export chart data as Excel format."""
    try:
        excel_b64 = _create_excel_with_openpyxl(chart, data, columns)
        return _create_excel_chart_data(
            chart, data, excel_b64, performance, cache_status
        )
    except ImportError:
        return _try_xlsxwriter_fallback(chart, data, columns, cache_status, performance)


def _create_excel_with_openpyxl(
    chart: "Slice", data: List[Dict[str, Any]], columns: List[str]
) -> str:
    """Create Excel file using openpyxl."""
    import base64
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = chart.slice_name[:31] if chart.slice_name else "Chart Data"

    if data and columns:
        _write_excel_headers(ws, columns)
        _write_excel_data(ws, data, columns)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return base64.b64encode(output.read()).decode()


def _write_excel_headers(ws: Any, columns: List[str]) -> None:
    """Write headers to Excel worksheet."""
    for idx, col in enumerate(columns, 1):
        ws.cell(row=1, column=idx, value=col)


def _write_excel_data(ws: Any, data: List[Dict[str, Any]], columns: List[str]) -> None:
    """Write data to Excel worksheet."""
    for row_idx, row in enumerate(data, 2):
        for col_idx, col in enumerate(columns, 1):
            value = row.get(col, "")
            if value is None:
                value = ""
            elif isinstance(value, (list, dict)):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)


def _try_xlsxwriter_fallback(
    chart: "Slice",
    data: List[Dict[str, Any]],
    columns: List[str],
    cache_status: Any,
    performance: Any,
) -> "ChartData | ChartError":
    """Try xlsxwriter as fallback for Excel export."""
    try:
        excel_b64 = _create_excel_with_xlsxwriter(chart, data, columns)
        return _create_excel_chart_data_xlsxwriter(
            chart, data, excel_b64, performance, cache_status
        )
    except ImportError:
        from superset.mcp_service.chart.schemas import ChartError

        return ChartError(
            error="Excel export requires openpyxl or xlsxwriter package",
            error_type="ExportError",
        )


def _create_excel_with_xlsxwriter(
    chart: "Slice", data: List[Dict[str, Any]], columns: List[str]
) -> str:
    """Create Excel file using xlsxwriter."""
    import base64
    import io

    import xlsxwriter

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    sheet_name = chart.slice_name[:31] if chart.slice_name else "Chart Data"
    worksheet = workbook.add_worksheet(sheet_name)

    if data and columns:
        _write_xlsxwriter_data(worksheet, data, columns)

    workbook.close()
    output.seek(0)
    return base64.b64encode(output.read()).decode()


def _write_xlsxwriter_data(
    worksheet: Any, data: List[Dict[str, Any]], columns: List[str]
) -> None:
    """Write data to xlsxwriter worksheet."""
    # Write headers
    for idx, col in enumerate(columns):
        worksheet.write(0, idx, col)

    # Write data
    for row_idx, row in enumerate(data):
        for col_idx, col in enumerate(columns):
            value = row.get(col, "")
            if value is None:
                value = ""
            elif isinstance(value, (list, dict)):
                value = str(value)
            worksheet.write(row_idx + 1, col_idx, value)


def _create_excel_chart_data(
    chart: "Slice",
    data: List[Dict[str, Any]],
    excel_b64: str,
    performance: Any,
    cache_status: Any,
) -> "ChartData":
    """Create ChartData response for Excel export (openpyxl)."""
    from superset.mcp_service.chart.schemas import ChartData

    chart_name = chart.slice_name or f"Chart {chart.id}"
    summary = f"Excel export of chart '{chart.slice_name}' with {len(data)} rows"

    return ChartData(
        chart_id=chart.id,
        chart_name=chart_name,
        chart_type=chart.viz_type or "unknown",
        columns=[],
        data=[],
        row_count=len(data),
        total_rows=len(data),
        summary=summary,
        insights=["Data exported as Excel format (base64 encoded)"],
        data_quality={},
        recommended_visualizations=[],
        data_freshness=None,
        performance=performance,
        cache_status=cache_status,
        excel_data=excel_b64,
        format="excel",
    )


def _create_excel_chart_data_xlsxwriter(
    chart: "Slice",
    data: List[Dict[str, Any]],
    excel_b64: str,
    performance: Any,
    cache_status: Any,
) -> "ChartData":
    """Create ChartData response for Excel export (xlsxwriter)."""
    from superset.mcp_service.chart.schemas import ChartData

    chart_name = chart.slice_name or f"Chart {chart.id}"
    summary = f"Excel export of chart '{chart.slice_name}' with {len(data)} rows"

    return ChartData(
        chart_id=chart.id,
        chart_name=chart_name,
        chart_type=chart.viz_type or "unknown",
        columns=[],
        data=[],
        row_count=len(data),
        total_rows=len(data),
        summary=summary,
        insights=["Data exported as Excel format (base64 encoded, xlsxwriter)"],
        data_quality={},
        recommended_visualizations=[],
        data_freshness=None,
        performance=performance,
        cache_status=cache_status,
        excel_data=excel_b64,
        format="excel",
    )
